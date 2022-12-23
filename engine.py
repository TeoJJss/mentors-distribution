import pandas as pd
from openpyxl import load_workbook
import numpy as np, os

def allocate(student, staff, i, j):
    # print(j, staff['emp_uniqueid'][j])
    student.loc[i, 'MENTOR']=staff['employeeno'][j]
    student.loc[i, 'SAMACCOUNTNAME'] = staff['emp_uniqueid'][j]
    student.loc[i, 'NAME'] = staff['emp_fullname'][j]
    staff.at[j, 'pax'] = staff['pax'][j]+1

files=os.listdir("input")
# files=['SoCNov22.xlsx']
for file in files:
    # time.sleep(7)
    file = "input/"+file
    print("\n\n", file)
    wb = load_workbook(file, read_only=True)
    #check if the required sheets in the input-file exists
    if 'Student' and 'Staff' in wb.sheetnames:
    # initialize
        student=pd.read_excel(file, 'Student').sort_values(by=['START_DATE'], ignore_index=True)
        staff = pd.read_excel(file, 'Staff')
        staff['Max pax'] = staff['Max pax'].fillna(0)
        # print(student)
        # print(student['MENTOR'].value_counts())
        # staff employeeno array
        code=[]
        for i in range(staff.shape[0]):
            code.append(staff.iloc[i]['employeeno'])

        # Assign current pax to staff
        count=[]
        for i in code: 
            count.append((student.MENTOR == i).sum())
        count=[int(i) for i in count]
        staff=staff.assign(pax=count)
        staff = staff.sort_values(by=['pax'], ignore_index=True)
        staff['pax'] = staff['pax'].fillna(0)

        j=0
        # print(staff['employeeno'][0])
        # Allocation Process
        for i in range(student.shape[0]): # Do until all students have mentors
        # Select staff
            #staff with max load of pax will be removed from dataframe
            staff['status']= np.where((staff['pax']<staff['Max pax']) | (staff['Max pax']==0.0), "active", "full") 
            staff=(staff[staff['status']=="active"]).reset_index(drop=True)

            # stop if all students have mentors
            if staff.empty | student['MENTOR'].notnull().all():
                break

            # Repeat change staff until the student get a mentor
            while pd.isna(student['MENTOR'][i]): 
                #Avoid out of range
                if j not in range(staff.shape[0]):
                    j=0
            
                # if all staffs have same number of pax, no condition to ignore
                if (staff['pax'] == staff['pax'][0]).all(): 
                    # Record in student dataframe
                    allocate(student, staff, i, j)
                    j+=1
                    
                    
                # if all staffs have diff. number of pax, max pax to ignore
                else: 
                    temp = staff[student]
                    rest = staff['pax'].mean(skipna=False) 
                    if staff['pax'][j] < rest and staff['pax'][j]!=staff['pax'].max(skipna=False) :     #ignore maximum number of pax for this cycle ONLY
                        # Record in student dataframe
                        allocate(student, staff, i, j)
                    j+=1


        # Prepare output file
        result = student.sort_values(by=['START_DATE', 'COURSE_CODE_ALIAS', 'TYPE_OF_COURSE', 'MENTOR'])
        # result = result[result['TYPE_OF_COURSE']=="Level 1"]
        print(result['MENTOR'].value_counts())
        print('output/result_'+(file.split("/")[-1]).split(".")[0]+".csv")
        result.to_csv('output/result_'+(file.split("/")[-1]).split(".")[0]+".csv", index=False)