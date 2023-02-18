# from cmath import isnan, nan
import pandas as pd
# import openpyxl as xl
from openpyxl import load_workbook
# import os, time, boto3
import numpy as np
# from flask import send_file

def allocate(student, staff, i, j):
    # print(j, staff['emp_uniqueid'][j])
    # print(i,j)
    student.loc[i, 'MENTOR']=staff['employeeno'][j]
    student.loc[i, 'SAMACCOUNTNAME'] = staff['emp_uniqueid'][j]
    student.loc[i, 'NAME'] = staff['emp_fullname'][j]
    staff.at[j, 'pax'] = staff['pax'][j]+1

# files=os.listdir("input")
def process(file):
    # # time.sleep(7)
    print(file.filename)
    # file = "input/"+file
    # print("\n\n", file)
    wb = load_workbook(file, read_only=True)
    #check if the required sheets in the input-file exists
    if 'Student' and 'Staff' in wb.sheetnames:
    # initialize
        student=pd.read_excel(file, 'Student').sort_values(by=['START_DATE'], ignore_index=True)
        staff = pd.read_excel(file, 'Staff')
        staff['Max pax'] = staff['Max pax'].fillna(0)
        # print(student)
        # print(student['MENTOR'].value_counts())
        # staff employeeno array
        print(staff.shape[0])
        code=[]
        for i in range(staff.shape[0]):
            code.append(staff.iloc[i]['employeeno'])

        # Assign current pax to staff
        count=[]
        for i in code: 
            count.append((student.MENTOR == i).sum())
        count=[int(i) for i in count]
        staff=staff.assign(pax=count)
        staff = staff.sort_values(by=['pax'], ignore_index=True)
        staff['pax'] = staff['pax'].fillna(0)
        staff['intake_pax']=""
        staff['level_pax']=""

        j=0
        # print(staff)
        # print(staff['employeeno'][0])
        # Allocation Process
        try:
            for i in range(student.shape[0]): # Do until all students have mentors
            # Select staff
                check=True
                #staff with max load of pax will be removed from dataframe
                # staff['status']= np.where((staff['pax']<staff['Max pax']) | (staff['Max pax']==0.0), "active", "full") 
                # staff=(staff[staff['status']=="active"]).reset_index(drop=True)

                # stop if all students have mentors
                if staff.empty | student['MENTOR'].notnull().all():
                    break

                # Repeat change staff until the student get a mentor
                while pd.isna(student['MENTOR'][i]): 
                    intake=student[student['START_DATE']==student['START_DATE'][i]]
                    level=student[student['TYPE_OF_COURSE']==student['TYPE_OF_COURSE'][i]]

                    # Identify each mentor's pax in each intake
                    intake_pax=[]
                    level_pax=[]
                    for k in range(staff.shape[0]):
                        intake_pax.append((intake['MENTOR']==staff['employeeno'][k]).sum())
                        level_pax.append((level['MENTOR']==staff['employeeno'][k]).sum())
                    staff['intake_pax']=intake_pax
                    staff['level_pax']=level_pax
                    # print(intake_pax)
                    #Avoid out of range
                    if j not in range(staff.shape[0]):
                        j=0
                    if (staff['intake_pax'] == staff['intake_pax'][0]).all() or (staff['level_pax'] == staff['level_pax'][0]).all():
                        check=False
                    
                    rest_intake=staff['intake_pax'].max(skipna=False)
                    rest_level=staff['level_pax'].max(skipna=False)
                    
                    if (staff['pax'] == staff['pax'][0]).all(): 
                        # Record in student dataframe
                        if check:
                            if staff['intake_pax'][j] < rest_intake or staff['level_pax'][j] < rest_level:
                                allocate(student, staff, i, j)
                        else:
                            allocate(student, staff, i, j)
                        j+=1
                        
                        
                    # if all staffs have diff. number of pax, max pax to ignore
                    else: 
                        
                        rest = staff['pax'].max(skipna=False) 
                        if check:
                            if staff['intake_pax'][j] < rest_intake or staff['level_pax'][j] < rest_level:  
                                allocate(student, staff, i, j)
                        else:
                            if staff['pax'][j] < rest:     #ignore maximum number of pax for this cycle ONLY
                                allocate(student, staff, i, j)
                        j+=1
        except:
            return student

        # Prepare output file
        result = student.sort_values(by=['START_DATE', 'TYPE_OF_COURSE', 'MENTOR'])
        print(result['MENTOR'].value_counts())
        
        print("\nL1 Sept")
        result2 = result[result['TYPE_OF_COURSE']=="Level 1"]
        result2 = result2[result2['START_DATE']=="2022-09-21"]
        # print(result)
        print(result2['MENTOR'].value_counts())
        # print('output/result_'+(file.filename).split(".")[0]+".csv")
        return result
        # result.to_csv('output/result_'+((file.filename).split("/")[-1]).split(".")[0]+".csv", index=False)
        # return send_file('output/result_'+((file.filename).split("/")[-1]).split(".")[0]+".csv" , mimetype='text/csv',  as_attachment=True)